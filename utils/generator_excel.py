import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
import os

def generar_excel(data, file_path, sheet_name="Reporte", headers=None, auto_fit=True):
    """
    Genera un archivo Excel con los datos proporcionados.

    :param data: Lista de listas con los datos a incluir en el Excel.
    :param file_path: Ruta donde se guardará el archivo Excel.
    :param sheet_name: Nombre de la hoja de cálculo.
    :param headers: Lista de encabezados para las columnas.
    :param auto_fit: Si True, ajusta automáticamente el ancho de las columnas.
    """
    # Crear un libro y una hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Si se proporcionan headers, se usan como títulos de columnas
    if headers:
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

    # Rellenar la hoja con los datos
    start_row = 2 if headers else 1  # Cambiar el número de fila de inicio según si hay encabezados
    for row_num, row_data in enumerate(data, start=start_row):
        # Extrae solo los valores del diccionario para esa fila
        valores = list(row_data.values())
        for col_num, cell_value in enumerate(valores, start=1):
            ws.cell(row=row_num, column=col_num, value=cell_value)

    # Auto-ajustar el ancho de las columnas
    if auto_fit:
        for col_num in range(1, ws.max_column + 1):
            max_length = 0
            column = get_column_letter(col_num)
            for cell in ws[column]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)  # Agregar un margen de 2
            ws.column_dimensions[column].width = adjusted_width

    # Guardar el archivo Excel
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    wb.save(file_path)
